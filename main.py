import os
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import smtplib
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from email.mime.text import MIMEText
import asyncio
import datetime
import aiohttp

load_dotenv()

async def fetch_data(session, url):
    async with session.get(url) as response:
        if response.status == 200:
            return await response.json()
        else:
            print("Ошибка при запросе данных:", response.status)

async def download_exchange_rates():
    load_dotenv()

    current_date = datetime.datetime.now()
    first_day_of_current_month = current_date.replace(day=1)
    thirty_days_ago = first_day_of_current_month - datetime.timedelta(days=30)
    current_date_today = current_date - datetime.timedelta(days=1)

    url_base = "https://iss.moex.com//iss/statistics/engines/futures/markets/indicativerates/securities.json"

    async with aiohttp.ClientSession() as session:
        tasks = []
        current_day = thirty_days_ago
        exchange_rates = []
        while current_day <= current_date_today:
            current_date_str = current_day.strftime('%Y-%m-%d')
            url = f"{url_base}?date={current_date_str}&till={current_date_str}"
            tasks.append(fetch_data(session, url))
            current_day += datetime.timedelta(days=1)

        responses = await asyncio.gather(*tasks)
        for response_data in responses:
            if response_data:
                usd_to_rub_data_day = [record for record in response_data['securities']['data'] if
                                       record[2] == 'USD/RUB']
                jpy_to_rub_data_day = [record for record in response_data['securities']['data'] if
                                       record[2] == 'JPY/RUB']
                for usd_record, jpy_record in zip(usd_to_rub_data_day, jpy_to_rub_data_day):
                    exchange_rates.append((usd_record[0], usd_record[3], usd_record[1], jpy_record[0], jpy_record[3], jpy_record[1]))

    return exchange_rates

def save_to_excel(exchange_rates):
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Дата USD/RUB'
    ws['B1'] = 'Курс USD/RUB'
    ws['C1'] = 'Время USD/RUB'

    ws['D1'] = 'Дата JPY/RUB'
    ws['E1'] = 'Курс JPY/RUB'
    ws['F1'] = 'Время JPY/RUB'

    row_idx = 2
    for exchange_rate in exchange_rates:
        ws[f'A{row_idx}'], ws[f'B{row_idx}'], ws[f'C{row_idx}'], ws[f'D{row_idx}'], ws[f'E{row_idx}'], ws[f'F{row_idx}'] = exchange_rate
        ws[f'G{row_idx}'] = exchange_rate[1] / exchange_rate[4]  # Курс USD/RUB / Курс JPY/RUB
        row_idx += 1

    financial_format = NamedStyle(name='financial', number_format='#,##0.00')
    wb.add_named_style(financial_format)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=7):
        for cell in row:
            cell.style = 'financial'

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save("currency_data.xlsx")

def send_email():
    email_sender = os.getenv("EMAIL_SENDER")
    email_receiver = os.getenv("EMAIL_RECEIVER")
    password = os.getenv("EMAIL_PASSWORD")

    msg = MIMEMultipart()
    msg['From'] = email_sender
    msg['To'] = email_receiver
    msg['Subject'] = 'Отчет о курсах валют'

    filename = "currency_data.xlsx"
    attachment = open(filename, "rb")

    part = MIMEBase('application', 'octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= " + filename)

    msg.attach(part)

    wb = load_workbook(filename)
    ws = wb.active
    num_rows = ws.max_row

    if num_rows == 1:
        row_word = "строка"
    elif num_rows % 10 == 1 and num_rows % 100 != 11:
        row_word = "строка"
    elif 2 <= num_rows % 10 <= 4 and (num_rows % 100 < 10 or num_rows % 100 >= 20):
        row_word = "строки"
    else:
        row_word = "строк"

    # Добавляем информацию о количестве строк в письмо
    body = f"В прикрепленном Excel файле {num_rows} {row_word}."
    msg.attach(MIMEText(body, 'plain'))

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(email_sender, password)
    text = msg.as_string()
    server.sendmail(email_sender, email_receiver, text)
    server.quit()

async def main():
    exchange_rates = await download_exchange_rates()
    save_to_excel(exchange_rates)
    send_email()

if __name__ == "__main__":
    asyncio.run(main())
